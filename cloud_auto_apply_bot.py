import json
import smtplib
import time
import random
import os
import openpyxl
import subprocess
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime

# ==========================================
# DUAL SENDER ACCOUNTS CONFIGURATION
# ==========================================
GMAIL_PASS_1 = os.environ.get('GMAIL_APP_PASS_1', 'sznpitgpdmhpkvee')
GMAIL_PASS_2 = os.environ.get('GMAIL_APP_PASS_2', 'djggjmmezbjxofbi')

ACCOUNTS = [
    {"email": "amin87.azman@gmail.com", "pass": GMAIL_PASS_1},
    {"email": "aminazman.inspection@gmail.com", "pass": GMAIL_PASS_2}
]

BATCH_LIMIT = 50           # Ultra-Safe 50 emails per session (~25 per account)
MIN_DELAY_SECONDS = 60     # Ultra-Safe mode (60s min delay)
MAX_DELAY_SECONDS = 120    # Ultra-Safe mode (120s max delay)

JSON_FILE = "contacts_data.json"
EXCEL_FILE = "Master_Subsea_Contacts_v2.xlsx"
CV_PDF_PATH = os.path.join("01_Offshore_Certificates_FULL", "CV - Muhammad Amin Azman (CSWIP 3.4U Inspection Engineer).pdf")
ZIP_DOCS_PATH = r"c:\Users\amin8\Desktop\AG Projects\AG-Offshore\All_Certificates_Amin_Azman.zip"
PROJECT_DIR = os.getcwd()

def generate_email_content(contact, sender_email):
    company = contact.get("company", "Subsea Team")
    pic_name = contact.get("pic_name", "")
    email = contact.get("email", "").lower()
    
    subject = f"CSWIP 3.4U Subsea Inspection Engineer / Data Recorder - Muhammad Amin Azman (Freelance / Contract / Permanent)"
    salut = f"Hi {pic_name}," if pic_name else f"Hi {company} Recruitment Team,"
    body = f"""{salut}

I am writing to express my strong interest in joining {company} for upcoming offshore campaigns, ad-hoc freelancing mobilizations, contract, or permanent positions as a CSWIP 3.4U Subsea Inspection Engineer / Data Recorder. I am 100% open for freelancing, contract, or permanent roles and willing to relocate worldwide.

With over 280+ offshore days across 10+ subsea campaigns (including PETRONAS, PTTEP, CHOC, and RINA Class surveys), I possess extensive hands-on expertise in subsea data acquisition, pipeline tracking (TSS 440/350 & MBES), Flooded Member Detection (Cobalt-60 & Impact Subsea FMD), and digital inspection suites (Sirrihatt, EdgeDVR, IDAMS).

Attached is my latest CV (PDF format). Full supporting certificate packages and editable DOCX formats are available immediately upon request.

I am 100% available for immediate worldwide offshore mobilization, open for freelancing / contract / permanent roles, and willing to relocate. I look forward to hearing from you soon regarding opportunities with {company}.

Best regards,

MUHAMMAD AMIN BIN AZMAN
CSWIP 3.4U Subsea Inspection Engineer / Data Recorder
Mobile / WhatsApp: +60125065516
Email: {sender_email}
Location: Penang, Malaysia (Point of Hire | Freelance, Contract & Permanent | Willing to Relocate Worldwide)

__________________________________________________
VALID OFFSHORE CREDENTIALS & CERTIFICATIONS SUMMARY:
- CSWIP 3.4U Subsea Inspection Controller (Cert No: 542968 | Valid Jul 2028)
- OPITO FOET / BOSIET with CA-EBS & Travel Safely By Boat (Valid Dec 2026)
- PETRONAS & OEUK Offshore Medical Examinations (Valid Sep 2026)
- ADNOC Offshore HSE Induction & Medical Clearances (Valid Oct 2027)
- PETRONAS Offshore Safety Passport (OSP) & Competency Cards (Valid Sep 2026 / Mar 2026)
- OPITO Basic H2S (9014) & Solas Marine Safety Certifications
- Mercury Awareness Certification
- Seaman’s Discharge Book / Card & International Passport (Valid Mar 2029)
- UTM B.Eng (Hons) Petroleum Engineering Degree (CGPA: 3.44)
"""
    return subject, body

def push_to_vercel():
    try:
        subprocess.run(["git", "add", "contacts_data.json", "Master_Subsea_Contacts_v2.xlsx"], cwd=PROJECT_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["git", "commit", "-m", "Real-time dual account CRM status update"], cwd=PROJECT_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["git", "push", "origin", "main"], cwd=PROJECT_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("     Synced live status to Vercel CRM!")
    except Exception as e:
        print(f"     Git sync error: {e}")

def send_via_smtp(sender_acc, target_email, subject, body):
    msg = MIMEMultipart()
    msg["From"] = f"Muhammad Amin Azman <{sender_acc['email']}>"
    msg["To"] = target_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # Attach ONLY 1 Single Print-ready CV PDF (No ZIP attachments to pass enterprise firewalls)
    if os.path.exists(CV_PDF_PATH):
        with open(CV_PDF_PATH, "rb") as cv_f:
            part = MIMEApplication(cv_f.read(), Name=os.path.basename(CV_PDF_PATH))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(CV_PDF_PATH)}"'
            msg.attach(part)

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender_acc["email"], sender_acc["pass"])
    server.sendmail(sender_acc["email"], target_email, msg.as_string())
    server.quit()

def run_dual_account_campaign():
    print("=== DUAL-ACCOUNT SAFE AUTO-APPLY CAMPAIGN INITIALIZED ===")
    
    with open(JSON_FILE, "r", encoding="utf-8") as f:
        contacts = json.load(f)

    # Filter unsent contacts (excluding MCS, Vantris, Alam Maritim & Bounced/Paused)
    unsent_contacts = [
        c for c in contacts 
        if str(c.get("sent")).upper() not in ["TRUE", "TRUE (SIMULATED)", "BOUNCED", "REDIRECTED", "PAUSED_MANUAL"] 
        and c.get("email") 
        and not any(ex in c.get("email", "").lower() or ex in c.get("company", "").lower() for ex in ["mcsoil", "vantris", "alam-maritim", "alam maritim"])
    ]

    print(f"Total Active Database Contacts: {len(contacts)}")
    print(f"Unsent Active Contacts Available: {len(unsent_contacts)}")
    print(f"Batch Limit Set To: {BATCH_LIMIT} emails tonight\n")

    if not unsent_contacts:
        print("All active contacts have already been emailed!")
        return

    batch_to_send = unsent_contacts[:BATCH_LIMIT]
    sent_count = 0

    for idx, contact in enumerate(batch_to_send, 1):
        target_email = contact["email"].strip()
        company = contact.get("company", "Company").encode('ascii', 'ignore').decode('ascii')
        
        # Alternate between ACCOUNTS (Round Robin) by default, but force aminazman.inspection@gmail.com for MCS & Vantris
        if any(domain in target_email.lower() for domain in ["mcsoil.com", "vantrisenergy.com"]):
            sender_acc = [a for a in ACCOUNTS if "aminazman.inspection" in a["email"]][0]
        else:
            sender_acc = ACCOUNTS[(idx - 1) % len(ACCOUNTS)]
        
        subject, body = generate_email_content(contact, sender_acc["email"])
        safe_subject = subject.encode('ascii', 'ignore').decode('ascii')

        print(f"[{idx}/{len(batch_to_send)}] Processing: {company} <{target_email}>")
        print(f"     Sender Account: {sender_acc['email']}")
        print(f"     Subject: {safe_subject}")

        try:
            send_via_smtp(sender_acc, target_email, subject, body)
            print(f"     SUCCESS: Email sent to {target_email} via {sender_acc['email']}")
            
            contact["sent"] = "TRUE"
            contact["sent_date"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            contact["status"] = "SENT / PENDING"
            contact["sender_email"] = sender_acc["email"]
            sent_count += 1

            with open(JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(contacts, f, indent=2)
            
            push_to_vercel()

            # Randomized long delay (12 to 25 mins)
            if idx < len(batch_to_send):
                delay = random.randint(MIN_DELAY_SECONDS, MAX_DELAY_SECONDS)
                print(f"     Sleeping for {delay} seconds (~{delay//60} mins) before next send...\n")
                time.sleep(delay)

        except Exception as e:
            print(f"     FAILED to send to {target_email}: {e}")

    # Sync Excel
    print("\nSynchronizing campaign status to Master Excel v2...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Master Contacts"]
        for row in range(2, ws.max_row + 1):
            cell_email = str(ws.cell(row=row, column=6).value or '').strip().lower()
            for c in contacts:
                if c.get("email") and c["email"].strip().lower() == cell_email and c.get("sent_date"):
                    ws.cell(row=row, column=1, value=c["status"])
                    ws.cell(row=row, column=11, value=f"TRUE ({c['sent_date']})")
                    break
        wb.save(EXCEL_FILE)
        print("Master Excel v2 updated successfully!")
    except Exception as e:
        print(f"Excel sync error: {e}")

    print(f"\n=== DUAL-ACCOUNT BATCH COMPLETE: Processed {sent_count} emails ===")

if __name__ == "__main__":
    run_dual_account_campaign()
